from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('Order.xlsx')
wb2 = Workbook()
ws = wb.active
ws2 = wb2.active
ws2.title = "Data"

col_data = []
full_data = []
gold_data = []
silver_data = []
yellow_data = []
rose_data = []
indexGold = 1
indexSilver = 1
indexRose = 1
style = ""
length = ""
ring_size = ""


ws2.append(["", "GOLD", "", ""])
ws2.append(["", "", "", ""])
for row in range(2, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        char = get_column_letter(col)
        col_data.append(ws[char + str(row)].value)

    if ("Gold" in str(col_data[25])):
        ws2.append([col_data[32], "Gold", "", ""])
        
    col_data = []
        
ws2.append(["", "", "", ""])
ws2.append(["", "SILVER", "", ""])
ws2.append(["", "", "", ""])

for row in range(2, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        char = get_column_letter(col)
        col_data.append(ws[char + str(row)].value)
   
    if("Silver" in str(col_data[25])):
        
        ws2.append([col_data[32], "Silver", "", ""])
        
    col_data = []

ws2.append(["", "", "", ""])
ws2.append(["", "ROSE", "", ""])
ws2.append(["", "", "", ""])

for row in range(2, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        char = get_column_letter(col)
        col_data.append(ws[char + str(row)].value)

    if("Rose" in str(col_data[25])):
    
        ws2.append([col_data[32], "Rose", "", ""])
    
    col_data = []

# ws2.append(["", "", "", ""])
# ws2.append(["", "YELLOW", "", ""])
# ws2.append(["", "", "", ""])

# for row in range(2, ws.max_row + 1):
#     for col in range(1, ws.max_column + 1):
#         char = get_column_letter(col)
#         col_data.append(ws[char + str(row)].value)

#     if("Yellow" in str(col_data[25])):
        
#         ws2.append([col_data[32], "Yellow", "", ""])
        
#     col_data = []
            

    
    # full_data.append(col_data)
    # col_data = []


wb2.save('new_data.xlsx')