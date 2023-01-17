from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('Order.xlsx')
wb2 = Workbook()
ws = wb.active
ws2 = wb2.active
ws2.title = "Data"

col_data = []
full_data = []
gold_data = []
silver_data = []
yellow_data = []
rose_data = []
indexGold = 1
indexSilver = 1
indexRose = 1
style = ""
length = ""
ring_size = ""

for row in range(2, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        char = get_column_letter(col)
        col_data.append(ws[char + str(row)].value)

    if ("Gold" in str(col_data[25])):
        gold_data.append(col_data)
        if("Style" in str(col_data[25])):
            if("Both" in str(col_data[25])):
                ws2.append([col_data[32], "Gold", "Both", "Yok"])
            elif("Single" in str(col_data[25])):
                ws2.append([col_data[32], "Gold", "Single", "Yok"])
            elif("Spider Web" in str(col_data[25])):
                ws2.append([col_data, "Gold", "Spider Web", "Yok"])

        elif("Length" in str(col_data[25])):
            if("17 inches" in str(col_data[25])):
                ws2.append([col_data[32], "Gold", "17 inç", ])

    elif("Silver" in str(col_data[25])):
        silver_data.append(col_data)

    elif("Rose" in str(col_data[25])):
        rose_data.append(col_data)

    else:
        yellow_data.append([col_data[25]])
    
    full_data.append(col_data)
    col_data = []


# print(non_data, '\n')
print(indexGold + indexSilver + indexRose)
wb2.save('new_data.xlsx')