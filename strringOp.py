from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import re
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog as fd
from tkinter.messagebox import showinfo
from datetime import datetime

# create the root window
root = tk.Tk()
root.title('Tkinter Open File Dialog')
root.resizable(False, False)
root.geometry('300x150')

filePath = "/Users/pc/Desktop"

def select_file():
    filetypes = (
        ('text files', '*.xlsx'),
        ('All files', '*.*')
    )

    filename = fd.askopenfilename(
        title='Bir Dosya Seçin',
        initialdir='/',
        filetypes=filetypes)
    
    showinfo(
        title='Dosya Düzenlendi. Seçilen Dosya:',
        message=filename
    )
    
    wb = load_workbook(filename)
    wb2 = Workbook()
    ws = wb.active
    ws2 = wb2.active
    ws2.title = "Data"
    col_data = []
    col_data_variations = []
    time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col)
            col_data.append(ws[char + str(row)].value)
            
        if("Color" in str(col_data[25]) or "Style" in str(col_data[25]) or "Length" in str(col_data[25]) or "Personalization" in str(col_data[25]) or "Ring size" in str(col_data[25])):
            out = re.sub(r'[^:,]+:', '', str(col_data[25]))
            col_data_variations = out.split(',')
            arr_len=len(col_data_variations)
            if(arr_len < 4):
                col_data_variations.append("Yok")
                col_data_variations.append("Yok")
                col_data_variations.append("Yok")
                col_data_variations.append("Yok")
                
        ws2.append([col_data[32],col_data_variations[0],col_data_variations[1],col_data_variations[2]])
        col_data=[]
        

    wb2.save("DKMN New Order.xlsx")
    
    
    


# open button
open_button = ttk.Button(
    root,
    text='Open a File',
    command=select_file
)

open_button.pack(expand=True)


# run the application
root.mainloop()



